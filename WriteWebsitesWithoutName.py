from openpyxl import load_workbook

filepath = "C:/Users/burak/Desktop/Excel_SQL/PROJE/LifeQuality/CompanyList.xlsx"
workbook = load_workbook(filepath)
worksheet = workbook["Companies(Copy)"]

companies_websites = """
https://www.linkedin.com/company/jti/
https://www.linkedin.com/company/gtt/
https://www.linkedin.com/company/abc/
"""

websites = companies_websites.strip().split('\n')

for row, website in enumerate(websites, start=1):
    worksheet.cell(row=row, column=1).hyperlink = website.strip()

workbook.save(filepath)
workbook.close()