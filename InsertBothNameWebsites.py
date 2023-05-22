import re
from openpyxl import load_workbook

filepath = "C:/Users/burak/Desktop/CompanyList.xlsx"
workbook = load_workbook(filepath)
worksheet = workbook["Companies"]

companies_websites = """
First Company https://www.linkedin.com/company/jti/
Second and Strong Company https://www.linkedin.com/company/gtt/
Third and Weak Company https://www.linkedin.com/company/abc/
"""

websites = companies_websites.strip().split('\n')

for row, website in enumerate(websites, start=1):
    match = re.search(r'(.*?)\s+(https?://\S+)', website)
    if match:
        company_name = match.group(1)
        company_website = match.group(2)
        worksheet.cell(row=row, column=1).hyperlink = company_website.strip()
        worksheet.cell(row=row, column=1).value = company_name.strip()

workbook.save(filepath)
workbook.close()
